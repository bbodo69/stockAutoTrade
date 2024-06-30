import os
import pandas as pd
import openpyxl

def ReadExcel(filePath, sheetName):

    if os.path.isfile(filePath):
        df = pd.read_excel(filePath, dtype=str, sheet_name=sheetName)
    return df

def readExcelToDataFrame(fileName, sheetName):

    if os.path.isfile(fileName):
        df = pd.read_excel(fileName, dtype=str, sheet_name=sheetName)
        return df

def saveDFToNewExcel(fileName, sheetName, df):

    if os.path.isfile(fileName):
        os.remove(fileName)
        wb = openpyxl.Workbook()
        wb.save(fileName)
    else:
        wb = openpyxl.Workbook()
        wb.save(fileName)
    with pd.ExcelWriter(fileName, mode='a') as writer:
        workBook = writer.book

        try:
            if sheetName in workBook.sheetnames:
                workBook.remove(workBook[sheetName])
        except:
            print("Worksheet does not exist")
        df.to_excel(writer, sheet_name=sheetName, index=False)

def saveDFToAppendExcel(fileName, sheetName, df):

    if "KOSPI" in fileName:
        fileName.replace('KOSPI', '1')

    directory = os.path.dirname(fileName)
    if not os.path.exists(directory):
        os.makedirs(directory)
    if not os.path.isfile(fileName):
        wb = openpyxl.Workbook()
        wb.save(fileName)
    with pd.ExcelWriter(fileName, mode='a') as writer:
        workBook = writer.book
        try:
            workBook.remove(workBook[sheetName])
        except:
            print("Worksheet does not exist")

    with pd.ExcelWriter(fileName, mode='a') as writer:
        df.to_excel(writer, sheet_name=sheetName, index=True)


def saveDFToJson(fileName, df):
    df.to_json(path_or_buf=fileName, orient="records")
